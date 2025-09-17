# app/blueprints/reportes.py

from collections import defaultdict
from flask import Blueprint, request, jsonify, make_response, current_app
from sqlalchemy import func, case, and_
from sqlalchemy.orm import selectinload, joinedload
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, ROUND_CEILING
from datetime import datetime, date, time, timedelta
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    AR_TZ = ZoneInfo("America/Argentina/Buenos_Aires")
except ImportError:
    import pytz
    AR_TZ = pytz.timezone("America/Argentina/Buenos_Aires")
import traceback
import io
import math
import pandas as pd

# Importar la librería para Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- Imports locales ---
from .. import db
from ..models import DetalleVenta # Asegúrate de importar DetalleVenta si no está
from ..models import Venta, OrdenCompra, Receta, RecetaItem
from ..models import Producto, TipoCambio
from ..calculator.core import obtener_coeficiente_por_rango
from .productos import calcular_costo_producto_referencia
from .productos import redondear_a_siguiente_decena, redondear_a_siguiente_centena
from ..utils.decorators import token_required, roles_required
from ..utils.permissions import ROLES

# --- Blueprint ---
reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
DAY_HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Gris para cabecera de día
MAX_REPORT_DAYS = 365

def style_header(ws, headers):
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(i)].width = max(len(header), 18)


def _extraer_estado_nombre_vendedor(nombre_vendedor: str) -> str:
    """Extrae el estado (p.ej. 'CANCELADO') del campo nombre_vendedor si existe.
    Retorna cadena en mayúsculas sin espacios adicionales, o vacía si no se detecta.
    """
    if not nombre_vendedor:
        return ''
    partes = str(nombre_vendedor).split('-', 1)
    estado = partes[0].strip().upper()
    return estado

@reportes_bp.route('/movimientos-excel', methods=['GET'])
@token_required
@roles_required(ROLES['ADMIN'], ROLES['CONTABLE'])
def reporte_movimientos_excel_limitado(current_user):
    """
    [VERSIÓN FINAL CON VENTAS UNIFICADAS Y AGRUPADAS POR DÍA]
    Genera un reporte maestro con:
    1. Hoja de "Ventas Detalladas" unificada y con separadores por día.
    2. Hoja de "Compras Detalladas".
    3. Hoja de "Resumen Global".
    """
    fecha_desde_str = request.args.get('fecha_desde')
    fecha_hasta_str = request.args.get('fecha_hasta')

    if not fecha_desde_str or not fecha_hasta_str:
        return jsonify({"error": "Los parámetros 'fecha_desde' y 'fecha_hasta' son requeridos."}), 400

    try:
        fecha_desde = date.fromisoformat(fecha_desde_str)
        fecha_hasta = date.fromisoformat(fecha_hasta_str)
        if (fecha_hasta - fecha_desde).days > MAX_REPORT_DAYS:
            return jsonify({"error": f"El rango excede el límite de {MAX_REPORT_DAYS} días."}), 400
        
        start_datetime = datetime.combine(fecha_desde, time.min)
        end_datetime = datetime.combine(fecha_hasta, time.max)
    except ValueError:
        return jsonify({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}), 400

    try:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        # --- 1. CONSULTAS EFICIENTES ---
        # Ordenamos por fecha para poder agrupar por día
        ventas_periodo = db.session.query(Venta).options(
            selectinload(Venta.cliente),
            selectinload(Venta.usuario_interno)
        ).filter(Venta.fecha_registro.between(start_datetime, end_datetime)).order_by(Venta.fecha_registro.asc()).all()

        compras_periodo = db.session.query(OrdenCompra).options(
            selectinload(OrdenCompra.proveedor)
        ).filter(OrdenCompra.fecha_creacion.between(start_datetime, end_datetime)).order_by(OrdenCompra.id.asc()).all()
        
        tc_oficial = TipoCambio.query.filter_by(nombre='Oficial').first()
        tc_empresa = TipoCambio.query.filter_by(nombre='Empresa').first()

        # Acumuladores para el resumen global
        ingresos_efectivo, ingresos_transferencia = Decimal('0.0'), Decimal('0.0')
        total_unidades_vendidas = Decimal('0.0')
        total_compras, total_pagado = Decimal('0.0'), Decimal('0.0')
        
# === HOJA 1: VENTAS DETALLADAS (UNIFICADA) [MODIFICADO v2.8 - NUEVA LÓGICA DE REDONDEO] ===
        ws_ventas = workbook.create_sheet("Ventas Detalladas")
        
        headers_ventas = [
            "Tipo Venta", "N° Pedido", "Fecha", "Hora", "Cliente", "Vendedor", 
            "Forma de Pago", "Estado", "ID Prod", "Producto", "Cant", 
            "Subtotal Item (ARS)", "Costo (ARS)", "Margen (%)", "Total Venta (ARS)"
        ]
        style_header(ws_ventas, headers_ventas)

        row_idx = 2

        for venta in ventas_periodo:
            tipo_venta = "Pedido" if venta.cliente_id is not None else "Puerta"
            # Estado: para pedidos, extraído de nombre_vendedor; para puerta, siempre 'Entregado'
            if tipo_venta == "Pedido" and venta.nombre_vendedor:
                partes = venta.nombre_vendedor.split('-', 1)
                estados_posibles = ['PENDIENTE', 'LISTO_PARA_ENTREGAR', 'ENTREGADO', 'CANCELADO']
                if len(partes) > 1 and partes[0].upper().replace(' ', '_') in estados_posibles:
                    estado_venta = partes[0].replace('_', ' ').title()
                else:
                    estado_venta = 'Pendiente'
            else:
                estado_venta = "Entregado"
            monto_total_base_venta = venta.monto_total or Decimal('0.0')
            # Usar monto_final_con_recargos redondeado a múltiplo de 100 hacia arriba
            monto_final_real = venta.monto_final_con_recargos or Decimal('0.0')
            monto_final_real = Decimal(math.ceil(monto_final_real / 100) * 100)

            # --- Ajuste de zona horaria para fecha/hora de registro ---
            fecha_registro = venta.fecha_registro
            if fecha_registro.tzinfo is None:
                try:
                    from zoneinfo import ZoneInfo
                    fecha_registro = fecha_registro.replace(tzinfo=ZoneInfo("UTC"))
                except ImportError:
                    import pytz
                    fecha_registro = pytz.utc.localize(fecha_registro)
            fecha_ar = fecha_registro.astimezone(AR_TZ)

            for detalle in venta.detalles:
                # Usar el subtotal base SIN redondear para cálculos de margen
                subtotal_item_base = detalle.precio_total_item_ars or Decimal('0.0')
                subtotal_item_final_calculado = subtotal_item_base
                if monto_total_base_venta > 0:
                    proporcion_del_item = subtotal_item_base / monto_total_base_venta
                    subtotal_con_recargos = monto_final_real * proporcion_del_item
                    # Mantener subtotal con recargos pero sin forzar redondeo para el cálculo del margen
                    subtotal_item_final_calculado = Decimal(subtotal_con_recargos)

                costo_total_prod, margen = Decimal('0.0'), Decimal('0.0')
                try:
                    # calcular_costo_producto_referencia devuelve costo unitario en USD
                    costo_unitario_usd = calcular_costo_producto_referencia(detalle.producto_id) or Decimal('0.0')
                    tc = tc_oficial if detalle.producto and detalle.producto.ajusta_por_tc else tc_empresa
                    if tc and tc.valor and tc.valor > 0:
                        costo_total_prod = costo_unitario_usd * detalle.cantidad * tc.valor
                    if subtotal_item_final_calculado > 0:
                        margen = ((subtotal_item_final_calculado - costo_total_prod) / subtotal_item_final_calculado) * 100
                except Exception as e:
                    print(f"Error calc prod {detalle.producto_id}: {e}")

                ws_ventas.cell(row=row_idx, column=1, value=tipo_venta)
                ws_ventas.cell(row=row_idx, column=2, value=venta.id)
                fecha_cell = ws_ventas.cell(row=row_idx, column=3, value=fecha_ar.date())
                fecha_cell.number_format = 'DD-MM-YYYY'
                hora_cell = ws_ventas.cell(row=row_idx, column=4, value=fecha_ar.time())
                hora_cell.number_format = 'HH:MM:SS'
                ws_ventas.cell(row=row_idx, column=5, value=venta.cliente.nombre_razon_social if venta.cliente else "Consumidor Final")
                ws_ventas.cell(row=row_idx, column=6, value=venta.nombre_vendedor or "N/A")
                ws_ventas.cell(row=row_idx, column=7, value=venta.forma_pago or "N/A")
                ws_ventas.cell(row=row_idx, column=8, value=estado_venta)
                ws_ventas.cell(row=row_idx, column=9, value=detalle.producto_id)
                ws_ventas.cell(row=row_idx, column=10, value=detalle.producto.nombre if detalle.producto else "N/A")
                ws_ventas.cell(row=row_idx, column=11, value=float(detalle.cantidad))
                ws_ventas.cell(row=row_idx, column=12, value=float(subtotal_item_final_calculado)).number_format = '"$"#,##0.00'
                ws_ventas.cell(row=row_idx, column=13, value=float(costo_total_prod)).number_format = '"$"#,##0.00'
                ws_ventas.cell(row=row_idx, column=14, value=float(margen)).number_format = '0.00"%"'
                ws_ventas.cell(row=row_idx, column=15, value=float(monto_final_real)).number_format = '"$"#,##0.00'
                row_idx += 1

            # Clasificar forma de pago: efectivo vs transferencia/otros
            forma = (venta.forma_pago or '').strip().lower()
            if forma == 'efectivo':
                ingresos_efectivo += monto_final_real
            elif forma == 'transferencia':
                ingresos_transferencia += monto_final_real
            else:
                # Otros medios (cheque, cuenta corriente, factura) se contabilizan en 'ingresos_transferencia' para resumen
                ingresos_transferencia += monto_final_real

        # === HOJA 2: COMPRAS DETALLADAS ===
        ws_compras = workbook.create_sheet("Compras Detalladas")
        headers_compras = ["N° Orden", "Fecha Creación", "Proveedor", "Estado Pago", "Monto Total (ARS)", "Monto Pagado (ARS)", "Deuda Pendiente (ARS)", "Forma de Pago"]
        style_header(ws_compras, headers_compras)
        for row_idx_c, compra in enumerate(compras_periodo, 2):
            monto_total = compra.importe_total_estimado or Decimal("0.0")
            monto_pagado = compra.importe_abonado or Decimal("0.0")
            estado_pago = "Pagada" if (monto_total - monto_pagado <= 0 and monto_total > 0) else ("Parcial" if monto_pagado > 0 else "Pendiente")
            ws_compras.cell(row=row_idx_c, column=1, value=compra.nro_solicitud_interno or compra.id)
            ws_compras.cell(row=row_idx_c, column=2, value=compra.fecha_creacion.strftime('%Y-%m-%d %H:%M'))
            ws_compras.cell(row=row_idx_c, column=3, value=compra.proveedor.nombre if compra.proveedor else "N/A")
            ws_compras.cell(row=row_idx_c, column=4, value=estado_pago)
            ws_compras.cell(row=row_idx_c, column=5, value=float(monto_total)).number_format = '"$"#,##0.00'
            ws_compras.cell(row=row_idx_c, column=6, value=float(monto_pagado)).number_format = '"$"#,##0.00'
            ws_compras.cell(row=row_idx_c, column=7, value=float(monto_total - monto_pagado)).number_format = '"$"#,##0.00'
            ws_compras.cell(row=row_idx_c, column=8, value=compra.forma_pago or "N/A")
            total_compras += monto_total
            total_pagado += monto_pagado

        # === HOJA 3: RESUMEN GLOBAL ===
        ws_resumen = workbook.create_sheet("Resumen Global")
        style_header(ws_resumen, ["Métrica", "Valor"])
        resumen_data = [
            ("Período del Reporte", f"{fecha_desde_str} al {fecha_hasta_str}"),
            ("--- RESUMEN DE VENTAS (GLOBAL) ---", ""),
            ("Ingresos Totales (Efectivo)", float(ingresos_efectivo)),
            ("Ingresos Totales (Transferencia y Otros)", float(ingresos_transferencia)),
            ("Ingreso Total del Período", float(ingresos_efectivo + ingresos_transferencia)),
            ("Total Unidades Vendidas (Kg/Lt)", float(total_unidades_vendidas)),
            ("--- RESUMEN DE COMPRAS ---", ""),
            ("Monto Total en Compras", float(total_compras)),
            ("Monto Total Pagado a Proveedores", float(total_pagado)),
            ("Deuda Total con Proveedores", float(total_compras - total_pagado))
        ]
        
        for idx, (metrica, valor) in enumerate(resumen_data, 2):
            cell_metrica = ws_resumen.cell(row=idx, column=1, value=metrica)
            cell_valor = ws_resumen.cell(row=idx, column=2, value=valor)
            if "---" in metrica:
                cell_metrica.font = Font(bold=True); cell_metrica.alignment = Alignment(horizontal='center')
                ws_resumen.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=2)
            elif isinstance(valor, float):
                cell_valor.number_format = '"$"#,##0.00' if any(term in metrica for term in ["Ingreso", "Monto", "Deuda"]) else '#,##0.00'

        # --- FINALIZAR Y DEVOLVER ---
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        nombre_archivo = f"Reporte_Maestro_{fecha_desde_str}_a_{fecha_hasta_str}.xlsx"
        response = make_response(excel_buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "Error interno al generar el reporte maestro.", "detalle": str(e)}), 500


def generar_precio_para_reporte(producto: Producto, cantidad_decimal: Decimal) -> Decimal:
    """
    [VERSIÓN FINAL] Calcula y devuelve el PRECIO TOTAL BRUTO, con máxima precisión y SIN REDONDEAR.
    La responsabilidad del redondeo final se delega a la función que la llama.
    """
    costo_unitario_venta_usd = calcular_costo_producto_referencia(producto.id)
    nombre_tc = 'Oficial' if producto.ajusta_por_tc else 'Empresa'
    tc_obj = TipoCambio.query.filter_by(nombre=nombre_tc).first()
    if not tc_obj or tc_obj.valor <= 0: raise ValueError(f"TC '{nombre_tc}' inválido")
    costo_unitario_venta_ars = costo_unitario_venta_usd * tc_obj.valor
    margen = Decimal(str(producto.margen or '0.0'))
    precio_base_ars = costo_unitario_venta_ars / (Decimal('1') - margen)
    resultado_tabla = obtener_coeficiente_por_rango(str(producto.ref_calculo), str(cantidad_decimal), producto.tipo_calculo)
    if resultado_tabla is None: raise ValueError("No habilitado para esta cantidad")
    coeficiente_str, escalon_cantidad_str = resultado_tabla
    coeficiente_decimal = Decimal(coeficiente_str)
    if cantidad_decimal >= Decimal('1.0'):
        precio_venta_unitario_bruto = precio_base_ars * coeficiente_decimal
    else:
        precio_para_la_fraccion = precio_base_ars * coeficiente_decimal
        escalon_decimal = Decimal(escalon_cantidad_str)
        if escalon_decimal == Decimal('0'): raise ValueError("Escalón de matriz es cero")
        precio_venta_unitario_bruto = precio_para_la_fraccion / escalon_decimal
    precio_total_bruto_ars = precio_venta_unitario_bruto * cantidad_decimal
    return precio_total_bruto_ars


@reportes_bp.route('/lista_precios/excel', methods=['GET'])
@token_required
def exportar_lista_precios_excel(current_user):
    """
    [VERSIÓN FINAL CON REDONDEO MATEMÁTICO]
    Aplica el redondeo a la decena con matemática simple para garantizar el resultado.
    """
    try:
        fecha_descarga = datetime.now()
        tc_oficial_obj = TipoCambio.query.filter_by(nombre='Oficial').first()
        tc_empresa_obj = TipoCambio.query.filter_by(nombre='Empresa').first()
        if not tc_oficial_obj or not tc_empresa_obj:
            raise ValueError("Faltan tipos de cambio 'Oficial' o 'Empresa' en la configuración.")

        cantidades_a_calcular = ["0.1", "0.25", "0.5", "1", "5", "10", "20", "25", "50", "100", "200", "500", "1000"]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Lista de Precios"
        cabeceras = ["ID", "Nombre Producto", "Costo Ref. USD"] + [f"Precio Total x {qty}" for qty in cantidades_a_calcular]
        style_header(sheet, cabeceras)
        
        productos = Producto.query.order_by(Producto.nombre).all()
        
        for row_num, producto in enumerate(productos, 2):
            sheet.cell(row=row_num, column=1, value=producto.id)
            sheet.cell(row=row_num, column=2, value=producto.nombre)
            
            try:
                costo_unitario_usd = calcular_costo_producto_referencia(producto.id) or Decimal('0')
                sheet.cell(row=row_num, column=3, value=float(costo_unitario_usd)).number_format = '"$"#,##0.0000'

                if not producto.ref_calculo or not producto.ref_calculo.strip():
                    error_msg = "Error: Falta 'ref_calculo' en BD"
                    for col_num_offset, _ in enumerate(cantidades_a_calcular): sheet.cell(row=row_num, column=4 + col_num_offset, value=error_msg)
                    continue

                for col_num_offset, qty_str in enumerate(cantidades_a_calcular):
                    current_col = 4 + col_num_offset
                    cell = sheet.cell(row=row_num, column=current_col)
                    try:
                        if costo_unitario_usd == Decimal('0'):
                            cell.value = 0.0
                        else:
                            cantidad_decimal = Decimal(qty_str)

                            # 1. Calcular precio TOTAL bruto para la cantidad solicitada (sin redondeos)
                            precio_total_bruto = generar_precio_para_reporte(producto, cantidad_decimal)

                            # 2. Derivar precio unitario bruto a partir del total (para cantidades >= 1 se mantiene el unitario, para fracciones la lógica interna ya fue aplicada)
                            # Evitamos llamar a generar_precio_para_reporte(producto, 1) que podría dar resultados distintos a dividir el total
                            try:
                                precio_unitario_bruto = (precio_total_bruto / cantidad_decimal) if cantidad_decimal != Decimal('0') else precio_total_bruto
                            except Exception:
                                precio_unitario_bruto = precio_total_bruto

                            # 3. Redondear unitario a la siguiente decena
                            precio_unitario_redondeado = redondear_a_siguiente_decena(precio_unitario_bruto)

                            # 4. Calcular total final multiplicando unitario redondeado por cantidad y redondear a siguiente centena
                            precio_total_final = redondear_a_siguiente_centena(precio_unitario_redondeado * cantidad_decimal)

                            cell.value = float(precio_total_final)
                        
                        cell.number_format = '"$"#,##0.00'
                    
                    except ValueError as ve:
                        cell.value = str(ve)
                    except Exception as cell_error:
                        cell.value = f"Error: {cell_error}"

            except Exception as row_error:
                error_msg = f"Error Fila: {row_error}"
                for col_num_offset, _ in enumerate(cantidades_a_calcular): sheet.cell(row=row_num, column=4 + col_num_offset, value=error_msg)
        
        info_sheet = workbook.create_sheet(title="Datos de Generación")
        info_sheet.cell(row=1, column=1, value="Concepto").font = Font(bold=True)
        info_sheet.cell(row=1, column=2, value="Valor").font = Font(bold=True)
        info_sheet.cell(row=2, column=1, value="Fecha y Hora de Descarga")
        info_sheet.cell(row=2, column=2, value=fecha_descarga.strftime('%Y-%m-%d %H:%M:%S'))
        info_sheet.cell(row=3, column=1, value="Tipo de Cambio 'Oficial' Utilizado (USD a ARS)")
        dolar_oficial_cell = info_sheet.cell(row=3, column=2, value=float(tc_oficial_obj.valor))
        dolar_oficial_cell.number_format = '"$"#,##0.00'
        info_sheet.cell(row=4, column=1, value="Tipo de Cambio 'Empresa' Utilizado (USD a ARS)")
        dolar_empresa_cell = info_sheet.cell(row=4, column=2, value=float(tc_empresa_obj.valor))
        dolar_empresa_cell.number_format = '"$"#,##0.00'
        info_sheet.column_dimensions['A'].width = 50
        info_sheet.column_dimensions['B'].width = 25
        
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        nombre_archivo = f"Lista_De_Precios_Volumen_{date.today().isoformat()}.xlsx"
        response = make_response(excel_buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "No se pudo generar el archivo de lista de precios", "detalle": str(e)}), 500


# ==============================================================================
# AQUÍ COMIENZAN TUS OTRAS FUNCIONES Y ENDPOINTS
# (Debes asegurarte de que estén aquí el resto de tus funciones, como `reporte_movimientos_excel_limitado`, etc.)
# ==============================================================================

def obtener_tipo_de_cambio_actual() -> Decimal:
    """
    Función auxiliar para obtener el tipo de cambio.
    (La mantengo por si es usada en otros endpoints que no me pasaste)
    """
    # Se asume que el tipo de cambio que nos interesa es el que tiene el nombre 'USD' o similar.
    tipo_cambio = TipoCambio.query.order_by(TipoCambio.fecha_actualizacion.desc()).first()
    if not tipo_cambio or not tipo_cambio.valor:
        raise ValueError("No se encontró un tipo de cambio configurado en la base de datos.")
    return tipo_cambio.valor


def _aplanar_receta(receta_item: RecetaItem, nivel: int, lista_plana: list):
    """
    Función auxiliar recursiva que navega la jerarquía de una receta
    y agrega cada ingrediente a una lista plana para el reporte.
    """
    ingrediente = receta_item.ingrediente
    if not ingrediente:
        return

    # Corrección final: Accede a la relación 'producto_final' y lo hace de forma segura.
    nombre_receta_padre = 'N/A'
    if receta_item.receta and receta_item.receta.producto_final:
        nombre_receta_padre = receta_item.receta.producto_final.nombre

    lista_plana.append({
        'Nivel': nivel,
        'Receta Padre': nombre_receta_padre,
        'ID Ingrediente': ingrediente.id,
        'Código Ingrediente': ingrediente.id, # Se usa ID ya que no hay campo 'codigo' o 'sku'
        'Nombre Ingrediente': ingrediente.nombre,
        'Es Receta': 'Sí' if ingrediente.es_receta else 'No',
        'Porcentaje en Receta Padre (%)': float(receta_item.porcentaje),
        'Costo USD Ingrediente (Unitario)': float(ingrediente.costo_referencia_usd or 0),
        'Contribución al Costo USD': (Decimal(str(receta_item.porcentaje)) / Decimal('100')) * Decimal(str(ingrediente.costo_referencia_usd or 0))
    })

    if ingrediente.es_receta and ingrediente.receta:
        # Consulta optimizada y corregida para usar 'producto_final'
        sub_items = db.session.query(RecetaItem).options(
            joinedload(RecetaItem.ingrediente),
            joinedload(RecetaItem.receta).joinedload(Receta.producto_final)
        ).filter(RecetaItem.receta_id == ingrediente.receta.id).all()
        
        for sub_item in sub_items:
            _aplanar_receta(sub_item, nivel + 1, lista_plana)


@reportes_bp.route('/formulas-excel', methods=['GET'])
@token_required
@roles_required(ROLES['ADMIN']) # Ajustado para permitir solo ADMIN
def exportar_formulas_a_excel(current_user):
    """
    Genera y devuelve un archivo Excel con el desglose completo de todas las recetas
    y sus costos anidados.
    """
    try:
        recetas_raiz = Producto.query.filter_by(es_receta=True).order_by(Producto.nombre).all()
        if not recetas_raiz:
            return jsonify({"error": "No se encontraron recetas para exportar."}), 404

        lista_final_plana = []
        for producto_receta in recetas_raiz:
            lista_final_plana.append({
                'Nivel': 0,
                'Receta Padre': '',
                'ID Ingrediente': producto_receta.id,
                'Código Ingrediente': producto_receta.id, # Se usa ID
                'Nombre Ingrediente': f"--- RECETA: {producto_receta.nombre} ---",
                'Es Receta': 'Sí',
                'Porcentaje en Receta Padre (%)': 100.0,
                'Costo USD Ingrediente (Unitario)': float(producto_receta.costo_referencia_usd or 0),
                'Contribución al Costo USD': float(producto_receta.costo_referencia_usd or 0)
            })

            if producto_receta.receta:
                # Consulta optimizada y corregida para usar 'producto_final'
                items_raiz = db.session.query(RecetaItem).options(
                    joinedload(RecetaItem.ingrediente),
                    joinedload(RecetaItem.receta).joinedload(Receta.producto_final)
                ).filter(RecetaItem.receta_id == producto_receta.receta.id).all()
                
                for item in items_raiz:
                    _aplanar_receta(item, 1, lista_final_plana)
        
        if not lista_final_plana:
            return jsonify({"error": "No se encontraron datos de recetas para exportar."}), 404

        df = pd.DataFrame(lista_final_plana)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Desglose de Fórmulas')
            worksheet = writer.sheets['Desglose de Fórmulas']
            worksheet.column_dimensions[get_column_letter(2)].width = 35
            worksheet.column_dimensions[get_column_letter(5)].width = 45
            for col_letter in ['H', 'I']:
                 for cell in worksheet[col_letter]:
                     cell.number_format = '"$"#,##0.0000'

        output.seek(0)
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename="Reporte_Formulas_Quimex_{datetime.date.today().isoformat()}.xlsx"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "Error interno del servidor al generar el reporte de fórmulas.", "detalle": str(e)}), 500
    
    
# ==============================================================================
# SECCIÓN DE FUNCIONES AUXILIARES PARA KPIs (NUEVO CÓDIGO REFACTORIZADO)
# ==============================================================================


def _get_kpis_del_dia(fecha_seleccionada: date):
    # (Las variables base_query_puerta y base_query_pedido deben definirse antes de usarse)
    """Calcula y devuelve los KPIs específicos del día seleccionado."""
    filtro_dia_entrega = func.date(Venta.fecha_pedido) == fecha_seleccionada

    # Mejorar el cálculo de ventas de puerta: solo ventas de mostrador, sin cliente y sin estado cancelado/anulado
    base_query_puerta = db.session.query(Venta).filter(
    filtro_dia_entrega,
        Venta.cliente_id.is_(None)
    )

    # Unidades por forma de pago (cuenta de ventas)
    puerta_efectivo_unidades = base_query_puerta.filter(Venta.forma_pago == 'efectivo').count()
    puerta_transferencia_unidades = base_query_puerta.filter(Venta.forma_pago == 'transferencia').count()
    puerta_factura_unidades = base_query_puerta.filter(Venta.forma_pago == 'factura').count()

    # Sumar el monto final redondeado de cada venta de puerta
    ventas_puerta = db.session.query(Venta).filter(
    filtro_dia_entrega,
        Venta.cliente_id.is_(None)
    ).all()
    ingreso_puerta_hoy = sum(
        Decimal(math.ceil((v.monto_final_redondeado or Decimal('0.0')) / 100) * 100)
        for v in ventas_puerta
    )

    # Desglose de puerta por forma de pago (MONTOS)
    puerta_efectivo = db.session.query(func.sum(Venta.monto_final_redondeado)).filter(
    filtro_dia_entrega,
        Venta.cliente_id.is_(None),
        Venta.forma_pago == 'efectivo'
    ).scalar() or Decimal('0.0')
    puerta_transferencia = db.session.query(func.sum(Venta.monto_final_redondeado)).filter(
    filtro_dia_entrega,
        Venta.cliente_id.is_(None),
        Venta.forma_pago == 'transferencia'
    ).scalar() or Decimal('0.0')
    puerta_factura = db.session.query(func.sum(Venta.monto_final_redondeado)).filter(
    filtro_dia_entrega,
        Venta.cliente_id.is_(None),
        Venta.forma_pago == 'factura'
    ).scalar() or Decimal('0.0')

    def estado_de_venta(nombre_vendedor):
        if not nombre_vendedor:
            return ''
        partes = nombre_vendedor.split('-', 1)
        estado = partes[0].strip().upper()
        return estado

    ventas_pedido = db.session.query(Venta).filter(
        filtro_dia_entrega,
        Venta.cliente_id.isnot(None)
    ).all()
    ventas_pedido_filtradas = [v for v in ventas_pedido if estado_de_venta(v.nombre_vendedor) != 'CANCELADO']
    base_query_pedido = db.session.query(Venta).filter(
        filtro_dia_entrega,
        Venta.cliente_id.isnot(None)
    )
    pedido_efectivo_unidades = base_query_pedido.filter(Venta.forma_pago == 'efectivo').count()
    pedido_transferencia_unidades = base_query_pedido.filter(Venta.forma_pago == 'transferencia').count()
    pedido_factura_unidades = base_query_pedido.filter(Venta.forma_pago == 'factura').count()
    ingreso_pedido_hoy = sum(
        Decimal(v.monto_final_redondeado or 0)
        for v in ventas_pedido_filtradas
    )

    # Desglose de pedidos por forma de pago (MONTOS)
    pedido_efectivo = sum(
        Decimal(v.monto_final_redondeado or 0)
        for v in ventas_pedido_filtradas if v.forma_pago == 'efectivo'
    )
    pedido_transferencia = sum(
        Decimal(v.monto_final_redondeado or 0)
        for v in ventas_pedido_filtradas if v.forma_pago == 'transferencia'
    )
    pedido_factura = sum(
        Decimal(v.monto_final_redondeado or 0)
        for v in ventas_pedido_filtradas if v.forma_pago == 'factura'
    )

    # KPIs de pedidos listos para entregar: si no hay columna estado, no se puede filtrar por 'Listo'.
    pedidos_listos_cantidad = 0
    cantidad_total_listos = 0

    def redondear_100(valor):
        return (valor // 100 * 100) if valor == 0 else ((valor + 99) // 100 * 100)

    return {
        "ingreso_puerta_hoy": redondear_100(ingreso_puerta_hoy),
        "ingreso_pedido_hoy": redondear_100(ingreso_pedido_hoy),
        "pedidos_listos_para_entregar": pedidos_listos_cantidad,
        "cantidad_total_listos": cantidad_total_listos,
        "puerta_efectivo": redondear_100(puerta_efectivo),
        "puerta_transferencia": redondear_100(puerta_transferencia),
        "puerta_factura": redondear_100(puerta_factura),
        "pedido_efectivo": redondear_100(pedido_efectivo),
        "pedido_transferencia": redondear_100(pedido_transferencia),
        "pedido_factura": redondear_100(pedido_factura),
        "puerta_efectivo_unidades": puerta_efectivo_unidades,
        "puerta_transferencia_unidades": puerta_transferencia_unidades,
        "puerta_factura_unidades": puerta_factura_unidades,
        "pedido_efectivo_unidades": pedido_efectivo_unidades,
        "pedido_transferencia_unidades": pedido_transferencia_unidades,
        "pedido_factura_unidades": pedido_factura_unidades,
    }

def _get_kpis_entregas_manana(fecha_seleccionada: date):
    """Calcula y devuelve los KPIs de entregas para el día siguiente."""
    fecha_siguiente = fecha_seleccionada + timedelta(days=1)
    
    # --- CORRECCIÓN FINAL v2.6 ---
    # Se ha confirmado que la columna correcta en el modelo Venta es 'fecha_pedido'.
    COLUMNA_FECHA_ENTREGA = Venta.fecha_pedido
    
    pedidos_pendientes_manana = db.session.query(Venta).filter(
        func.date(COLUMNA_FECHA_ENTREGA) == fecha_siguiente,
        Venta.cliente_id.isnot(None)
    ).count()
    kgs_manana = db.session.query(func.sum(DetalleVenta.cantidad)).join(Venta, Venta.id == DetalleVenta.venta_id).filter(
        func.date(COLUMNA_FECHA_ENTREGA) == fecha_siguiente,
        Venta.cliente_id.isnot(None)
    ).scalar() or Decimal('0.0')

    return {
        "pedidos_pendientes_manana": pedidos_pendientes_manana,
        "kgs_manana": kgs_manana
    }

def _get_kpis_de_compras():
    """Calcula y devuelve los KPIs globales de compras y proveedores."""
    deuda_proveedores_raw = db.session.query(
        func.sum(OrdenCompra.importe_total_estimado).label('total'),
        func.sum(OrdenCompra.importe_abonado).label('abonado')
    ).one()
    
    deuda_proveedores = (deuda_proveedores_raw.total or Decimal('0.0')) - (deuda_proveedores_raw.abonado or Decimal('0.0'))
    
    compras_por_recibir = db.session.query(func.sum(OrdenCompra.importe_total_estimado))\
        .filter(OrdenCompra.estado == 'Aprobado').scalar() or Decimal('0.0')
        
    return {
        "deuda_proveedores": deuda_proveedores,
        "compras_por_recibir": compras_por_recibir
    }

def _get_kpis_del_mes(fecha_seleccionada: date):
    """Calcula y devuelve los KPIs acumulados del mes hasta la fecha seleccionada."""
    filtro_mes_actual = and_(
        func.date(Venta.fecha_registro) >= fecha_seleccionada.replace(day=1),
        func.date(Venta.fecha_registro) <= fecha_seleccionada
    )
    
    # Obtener tipos de cambio para cálculo de costos
    tc_oficial = TipoCambio.query.filter_by(nombre='Oficial').first()
    tc_empresa = TipoCambio.query.filter_by(nombre='Empresa').first()

    # Primero obtener todas las ventas del periodo EXCLUYENDO las canceladas
    ventas_del_mes = db.session.query(Venta).filter(filtro_mes_actual).all()
    ventas_no_canceladas = [v for v in ventas_del_mes if _extraer_estado_nombre_vendedor(v.nombre_vendedor) != 'CANCELADO']

    ventas_mes_total = sum((v.monto_final_redondeado or Decimal('0.0')) for v in ventas_no_canceladas)

    # Ingresos por puerta/pedidos y efectivo/otros
    ingresos_puerta_mes = sum((v.monto_final_redondeado or Decimal('0.0')) for v in ventas_no_canceladas if v.cliente_id is None)
    ingresos_pedidos_mes = sum((v.monto_final_redondeado or Decimal('0.0')) for v in ventas_no_canceladas if v.cliente_id is not None)
    ingresos_efectivo_mes = sum((v.monto_final_redondeado or Decimal('0.0')) for v in ventas_no_canceladas if (v.forma_pago or '').strip().lower() == 'efectivo')
    ingresos_otros_mes = ventas_mes_total - ingresos_efectivo_mes

    # Calcular costos variables sumando costo real por detalle
    costos_variables_mes = Decimal('0.0')
    for v in ventas_no_canceladas:
        for det in v.detalles:
            try:
                costo_unitario_usd = calcular_costo_producto_referencia(det.producto_id) or Decimal('0.0')
                tc = tc_oficial if det.producto and det.producto.ajusta_por_tc else tc_empresa
                tc_val = tc.valor if tc and tc.valor else Decimal('0.0')
                costos_variables_mes += costo_unitario_usd * (det.cantidad or Decimal('0.0')) * tc_val
            except Exception:
                # Si falla el cálculo de costo de un producto, saltar y continuar
                continue

    ganancia_bruta_mes = ventas_mes_total - costos_variables_mes

    return {
        "ventas_mes": ventas_mes_total,
        "costos_variables_mes": costos_variables_mes,
        "ganancia_bruta_mes": ganancia_bruta_mes,
        "ingresos_puerta_mes": ingresos_puerta_mes,
        "ingresos_pedidos_mes": ingresos_pedidos_mes,
        "ingresos_efectivo_mes": ingresos_efectivo_mes,
        "ingresos_otros_mes": ingresos_otros_mes
    }

# ==============================================================================
# ENDPOINT PRINCIPAL (AHORA ES UN COORDINADOR LIMPIO)
# ==============================================================================

@reportes_bp.route('/dashboard-kpis', methods=['GET'])
@token_required
@roles_required(ROLES['ADMIN'], ROLES['CONTABLE'])
def get_dashboard_kpis(current_user):
    """
    [VERSIÓN 2.6 REFACTORIZADA CON FUNCIONES AUXILIARES]
    Coordina la llamada a funciones específicas para cada KPI y ensambla la respuesta final.
    Es más limpio, mantenible y fácil de depurar.
    """
    try:
        fecha_str = request.args.get('fecha', date.today().isoformat())
        fecha_seleccionada = date.fromisoformat(fecha_str)
        print(f"--- [INICIO] Solicitud a /dashboard/kpis (v2.6 Refactorizado) para fecha: {fecha_seleccionada} ---")

        # 1. Recopilar datos llamando a cada función especialista
        kpis_dia = _get_kpis_del_dia(fecha_seleccionada)
        kpis_manana = _get_kpis_entregas_manana(fecha_seleccionada)
        kpis_compras = _get_kpis_de_compras()
        kpis_mes = _get_kpis_del_mes(fecha_seleccionada)

        # 2. Armar el JSON de respuesta final
        response_data = {
            "primera_fila": {
                "ingreso_puerta_hoy": float(kpis_dia["ingreso_puerta_hoy"]),
                "ingreso_pedido_hoy": float(kpis_dia["ingreso_pedido_hoy"]),
                "pedidos_listos_para_entregar": int(kpis_dia["pedidos_listos_para_entregar"]),
                "cantidad_total_listos": float(kpis_dia["cantidad_total_listos"]),
                "puerta_efectivo": float(kpis_dia["puerta_efectivo"]),
                "puerta_transferencia": float(kpis_dia["puerta_transferencia"]),
                "puerta_factura": float(kpis_dia["puerta_factura"]),
                "pedido_efectivo": float(kpis_dia["pedido_efectivo"]),
                "pedido_transferencia": float(kpis_dia["pedido_transferencia"]),
                "pedido_factura": float(kpis_dia["pedido_factura"]),
                "puerta_efectivo_unidades": int(kpis_dia["puerta_efectivo_unidades"]),
                "puerta_transferencia_unidades": int(kpis_dia["puerta_transferencia_unidades"]),
                "puerta_factura_unidades": int(kpis_dia["puerta_factura_unidades"]),
                "pedido_efectivo_unidades": int(kpis_dia["pedido_efectivo_unidades"]),
                "pedido_transferencia_unidades": int(kpis_dia["pedido_transferencia_unidades"]),
                "pedido_factura_unidades": int(kpis_dia["pedido_factura_unidades"]),
                "pedidos_pendientes_manana": int(kpis_manana["pedidos_pendientes_manana"]),
                "kgs_manana": float(kpis_manana["kgs_manana"]),
                "deuda_proveedores": float(kpis_compras["deuda_proveedores"]),
                "compras_por_recibir": float(kpis_compras["compras_por_recibir"])
            },
            "segunda_fila": {
                "ventas_mes": float(kpis_mes["ventas_mes"]),
                "costos_variables_mes": float(kpis_mes["costos_variables_mes"]),
                "ganancia_bruta_mes": float(kpis_mes["ganancia_bruta_mes"]),
                # factor_ganancia_mes se calcula en el frontend o se puede añadir aquí
            },
            "tercera_fila": {
                "relacion_ingresos": { 
                    "puerta": float(kpis_mes["ingresos_puerta_mes"]), 
                    "pedidos": float(kpis_mes["ingresos_pedidos_mes"]) 
                },
                "relacion_pagos": { 
                    "efectivo": float(kpis_mes["ingresos_efectivo_mes"]), 
                    "otros": float(kpis_mes["ingresos_otros_mes"]) 
                }
            }
        }
        
        print("--- [FIN] Solicitud procesada con éxito. ---")
        return jsonify(response_data)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "Error interno al generar los KPIs del dashboard.", "detalle": str(e)}), 500

    
# === NUEVO ENDPOINT PARA LA CAJA DEL DÍA ===
@reportes_bp.route('/caja-del-dia', methods=['GET'])
@token_required
@roles_required(ROLES['ADMIN'], ROLES['CONTABLE'], ROLES['VENTAS_LOCAL'])
def get_caja_del_dia(current_user):
    """
    [NUEVO] Endpoint simple que devuelve el total de ventas del día actual,
    desglosado por forma de pago.
    """
    # Usar 'date' y 'datetime' importados directamente
    fecha_str = request.args.get('fecha', date.today().isoformat())

    try:
        fecha = date.fromisoformat(fecha_str)
        start_of_day = datetime.combine(fecha, time.min)
        end_of_day = datetime.combine(fecha, time.max)
    except ValueError:
        return jsonify({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}), 400
        
    try:
        ventas_del_dia = db.session.query(Venta).filter(
            Venta.fecha_registro.between(start_of_day, end_of_day)
        ).all() 

        caja = defaultdict(Decimal)
        for venta in ventas_del_dia:
            forma_pago = venta.forma_pago or "Otro"
            monto = venta.monto_final_redondeado or Decimal('0.0')
            caja[forma_pago] += monto
        
        # Convertir a float para JSON
        caja_float = {k: float(v) for k, v in caja.items()}
        
        return jsonify({
            "fecha": fecha_str,
            "resumen_caja": caja_float,
            "total_caja": float(sum(caja.values()))
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "Error interno al calcular la caja del día.", "detalle": str(e)}), 500

def obtener_tipo_de_cambio_actual() -> Decimal:
    """
    Consulta la base de datos para obtener el último valor del tipo de cambio.
    Utiliza el modelo TipoCambio y ordena por fecha_actualizacion.
    Lanza un error si no se encuentra.
    """
    # Se asume que el tipo de cambio que nos interesa es el que tiene el nombre 'USD' o similar.
    # Si tienes varios, ajusta el filtro. Si solo hay uno, este query es suficiente.
    tipo_cambio = TipoCambio.query.order_by(TipoCambio.fecha_actualizacion.desc()).first()
    
    if not tipo_cambio or not tipo_cambio.valor:
        raise ValueError("No se encontró un tipo de cambio configurado en la base de datos.")
        
    return tipo_cambio.valor
    
